from __future__ import unicode_literals
from django.shortcuts import render,redirect,HttpResponse
from django import forms
from django.db import connection
import django_excel as excel
import pyodbc
import openpyxl
from django.contrib import messages
from openpyxl import load_workbook, workbook
from .models import Parts 


secret_key = 'TARDIS' 

cursor = connection.cursor()

def index(request):
  
    return render(request,'index.html') 

def summ(request):
    
    return render(request, 'summ.html')

def inputData(request):
    
    return render(request, 'input_data.html')

def brows(request):
    
    return render(request, 'brows.html')

def pn(request):
    
    return render(request, 'lookuppn.html')

def desc(request):
    
    return render(request, 'lookupdesc.html')

def bulk(request):
    
    return render(request, 'lookupbulk.html')

excel_data = list()
def uploadData(request):
  
    if request.method == "POST":
        excel_file = request.FILES['excel_file']
        wb=openpyxl.load_workbook(excel_file)
        worksheet=wb['Sheet1']
    
        # for row in worksheet.iter_rows():
        #     row_data = list()
        #     # thisRow = list()
        #     for cell in row:              
        #         row_data.append(str(cell.value))

        #     excel_data.append(row_data)
        # row_num=2
        # for row_num in worksheet.iter_rows():
        row_data = list()
        part_num=worksheet.cell(row=2, column=1)
        description=worksheet.cell(row=2, column=2)
        row_data.append(str(part_num))
        row_data.append(str(description))





        excel_data.append(row_data)
        context={"excel_data":excel_data}

        return redirect('/bulksearch', context)
    else:
        return render(request,'lookupbulk.html', {})

""" def show(request, str):
    excel_file = request.FILES[str]
    wb=openpyxl.load_workbook(excel_file)
    worksheet=wb['Sheet1']
    excel_data = list()
    for row in worksheet.iter_rows():
        row_data = list()
       # thisRow = list()
        for cell in row:
            if cell.value.upper() =='PN':
                continue
            else:
                
                row_data.append(str(cell.value))

        excel_data.append(row_data) """


def bulksearch(request):
    wb=openpyxl.Workbook()
    ws = wb.active
    counter=1
    i=1
    for row in ws.iter_rows():
        for cell in row:
            if len(excel_data[counter]) >0:

                cell.value='A' #str(excel_data[counter])
                # print(excel_data[counter])
                counter +=1
        i += 1
    
    wb.save(filename = 'temp.xlsx')
    # str='temp.xlsx'
    # show(str)
    return render(request,'showdata.html')

    

   